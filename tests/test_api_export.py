"""Tests for GET /api/export.xlsx — full-catalog workbook download."""
from fastapi.testclient import TestClient

from ndif_citations.server import deps
from ndif_citations.server.app import create_app


def _client(fixture_state):
    app = create_app()
    app.dependency_overrides[deps.get_output_dir] = lambda: fixture_state
    return TestClient(app)


def test_export_xlsx_downloads_workbook(fixture_state):
    r = _client(fixture_state).get("/api/export.xlsx")
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers.get("content-disposition", "")
    # body is a valid xlsx with the expected sheets
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "GitHub" in wb.sheetnames
    assert "Papers" in wb.sheetnames
