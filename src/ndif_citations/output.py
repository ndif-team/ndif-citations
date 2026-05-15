"""Phase 4: Output — JSON/CSV writing with append/merge logic and CLI report."""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime
from pathlib import Path

from ndif_citations.models import Bucket, Category, DiscoveredPaper, DiscoveredRepo, PipelineRun
from ndif_citations.utils import is_duplicate

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Load existing data
# ---------------------------------------------------------------------------

def load_existing_papers(output_dir: Path) -> list[DiscoveredPaper]:
    """Load existing papers from the full JSON file.

    Expects the 3-bucket structure: {"pending": [...], "verified": [...], "discarded": [...]}.
    Raises ValueError if the file uses the old flat-list format.
    """
    full_json = output_dir / "research-papers-full.json"
    if not full_json.exists():
        return []

    try:
        with open(full_json) as f:
            data = json.load(f)

        if isinstance(data, list):
            raise ValueError(
                f"{full_json} uses the old flat-list format. "
                "Delete it or run with --fresh to regenerate in the new 3-bucket format: "
                '{"pending": [...], "verified": [...], "discarded": [...]}'
            )

        if not isinstance(data, dict) or not {"pending", "verified", "discarded"}.issubset(data.keys()):
            raise ValueError(
                f"{full_json} is missing required top-level keys. "
                'Expected {"pending": [...], "verified": [...], "discarded": [...]}'
            )

        papers: list[DiscoveredPaper] = []
        for bucket_key in ("pending", "verified", "discarded"):
            for item in data[bucket_key]:
                try:
                    paper = DiscoveredPaper.model_validate(item)
                    papers.append(paper)
                except Exception as e:
                    logger.warning(f"Failed to parse existing paper: {e}")

        logger.info(f"Loaded {len(papers)} existing papers from {full_json}")
        return papers

    except ValueError:
        raise
    except Exception as e:
        logger.warning(f"Failed to load existing papers: {e}")
        return []


def load_existing_repos(output_dir: Path) -> list[DiscoveredRepo]:
    """Load existing repos from the full JSON file."""
    full_json = output_dir / "github-repos-full.json"
    if not full_json.exists():
        return []

    try:
        with open(full_json) as f:
            data = json.load(f)

        repos = []
        for item in data:
            try:
                repo = DiscoveredRepo.model_validate(item)
                repos.append(repo)
            except Exception as e:
                logger.warning(f"Failed to parse existing repo: {e}")

        logger.info(f"Loaded {len(repos)} existing repos from {full_json}")
        return repos

    except Exception as e:
        logger.warning(f"Failed to load existing repos: {e}")
        return []


# ---------------------------------------------------------------------------
# Merge logic
# ---------------------------------------------------------------------------

def merge_papers(existing: list[DiscoveredPaper],
                 new: list[DiscoveredPaper],
                 run: PipelineRun | None = None) -> tuple[list[DiscoveredPaper], PipelineRun]:
    """Merge new papers into existing data.

    Returns (merged_list, run_stats).
    """
    if run is None:
        run = PipelineRun()

    if not existing:
        # First run -- everything is new
        run.new_papers = len(new)
        run.total_unique = len(new)
        return new, run

    # Build lookup from existing papers
    existing_by_arxiv: dict[str, int] = {}
    existing_by_doi: dict[str, int] = {}
    existing_titles: list[tuple[int, str]] = []

    for i, p in enumerate(existing):
        if p.arxiv_id:
            existing_by_arxiv[p.arxiv_id] = i
        if p.doi:
            existing_by_doi[p.doi] = i
        existing_titles.append((i, p.title))

    merged = list(existing)  # Start with existing
    new_count = 0
    updated_count = 0

    for paper in new:
        match_idx: int | None = None

        # Check arXiv ID
        if paper.arxiv_id and paper.arxiv_id in existing_by_arxiv:
            match_idx = existing_by_arxiv[paper.arxiv_id]

        # Check DOI
        if match_idx is None and paper.doi and paper.doi in existing_by_doi:
            match_idx = existing_by_doi[paper.doi]

        # Check title similarity
        if match_idx is None:
            for idx, title in existing_titles:
                if is_duplicate(paper.title, title):
                    match_idx = idx
                    break

        if match_idx is not None:
            # Paper already exists -- update non-manual fields
            existing_paper = merged[match_idx]
            was_updated = _update_existing(existing_paper, paper)
            if was_updated:
                updated_count += 1

            # Auto-recovery / auto-demotion: re-evaluate bucket unless manual_override
            if not existing_paper.manual_override:
                from ndif_citations.process import _decide_bucket
                new_bucket, new_reason = _decide_bucket(existing_paper)
                old_bucket = existing_paper.bucket
                if new_bucket != old_bucket:
                    if new_bucket == Bucket.VERIFIED and old_bucket == Bucket.PENDING:
                        run.auto_promoted.append(existing_paper.title)
                        logger.info(f"Auto-promoted: '{existing_paper.title[:60]}' (was: {old_bucket.value} {existing_paper.reason})")
                    elif new_bucket == Bucket.PENDING and old_bucket == Bucket.VERIFIED:
                        run.auto_demoted.append(existing_paper.title)
                        logger.info(f"Auto-demoted: '{existing_paper.title[:60]}' → {new_reason}")
                    existing_paper.bucket = new_bucket
                    existing_paper.reason = new_reason
                    if new_bucket == Bucket.VERIFIED:
                        existing_paper.reason_detail = None
        else:
            # New paper!
            merged.append(paper)
            new_count += 1
            # Update lookups
            idx = len(merged) - 1
            if paper.arxiv_id:
                existing_by_arxiv[paper.arxiv_id] = idx
            if paper.doi:
                existing_by_doi[paper.doi] = idx
            existing_titles.append((idx, paper.title))

    run.existing_papers = len(existing)
    run.new_papers = new_count
    run.updated_papers = updated_count
    run.total_unique = len(merged)

    logger.info(f"Merge: {new_count} new, {updated_count} updated, {len(existing)} existing -> {len(merged)} total")
    return merged, run


def merge_repos(
    discovered: list[DiscoveredRepo],
    existing: list[DiscoveredRepo],
) -> list[DiscoveredRepo]:
    """Merge discovered repos into existing state.

    - Stale entries (404 / renamed / archived) are NOT in discovered list
      (already dropped by enrich_repos_from_github_api). This function
      therefore purges any existing entry that doesn't appear in discovered,
      UNLESS it has manual_override=True.
    - For repos that appear in both: carry over existing state for SKIP/PROTECTED;
      update fields for NEW/REPROCESS/FILL_GAPS (routing has already set processing_bucket).
    - New repos (in discovered but not in existing): append as-is.
    """
    by_key: dict[str, DiscoveredRepo] = {r.merge_key(): r for r in existing}
    discovered_keys: set[str] = {r.merge_key() for r in discovered}

    merged: list[DiscoveredRepo] = []

    # Keep existing repos that are still discovered (or protected)
    removed_count = 0
    for existing_repo in existing:
        key = existing_repo.merge_key()
        if key in discovered_keys:
            continue  # Will be handled in the discovered loop below
        if existing_repo.manual_override:
            logger.debug(f"Keeping protected repo not seen this run: {key}")
            merged.append(existing_repo)
        else:
            logger.info(f"Purging stale repo from state: {key}")
            removed_count += 1

    if removed_count:
        logger.info(f"Purged {removed_count} stale repo(s) from github-repos-full.json")

    # Merge each discovered repo
    for repo in discovered:
        key = repo.merge_key()
        existing_repo = by_key.get(key)

        if existing_repo is None:
            # New repo — append as-is
            merged.append(repo)
        elif repo.processing_bucket in ("skip", "protected"):
            # Keep existing state (routing already swapped in the existing object in process_repos,
            # but guard here too)
            merged.append(existing_repo)
        else:
            # NEW / REPROCESS / FILL_GAPS — use freshly processed repo
            # Preserve manual_override from existing
            if existing_repo.manual_override:
                repo.manual_override = True
            merged.append(repo)

    return merged


def _update_existing(existing: DiscoveredPaper, new: DiscoveredPaper) -> bool:
    """Update an existing paper with new data, preserving manual overrides.

    Returns True if anything was changed.
    """
    changed = False

    # Never overwrite manually edited fields
    if existing.manual_override:
        return False

    # Update API-sourced fields if they've improved
    if new.authors and len(new.authors) > len(existing.authors):
        existing.authors = new.authors
        changed = True

    if new.affiliations and not existing.affiliations:
        existing.affiliations = new.affiliations
        changed = True

    # Venue: prefer the freshly-enriched (post-resolve_venue) value, but DO NOT
    # downgrade a CONFIDENTLY-RECOGNIZED existing venue to a preprint fallback.
    # The merge cases we care about:
    #   1. new=ArXiv fallback, existing=confident (e.g. "ICML 2025") → keep existing,
    #      re-normalize so any stale long-form gets cleaned up
    #   2. new=ArXiv fallback, existing=junk/truncated ("Handbook of Human 2025") →
    #      take new — ArXiv is cleaner than an unrecognized stub
    #   3. new=confident, anything → take new (it's the latest authoritative resolution)
    from ndif_citations.extract import detect_venue_type
    from ndif_citations.venue import (
        is_confident_venue, is_preprint_sentinel, normalize_venue,
    )
    if new.venue and new.venue != existing.venue:
        new_is_fallback = is_preprint_sentinel(new.venue)
        existing_is_confident = is_confident_venue(existing.venue)
        if new_is_fallback and existing_is_confident:
            # Case 1: don't downgrade — re-normalize existing in case it was
            # a pre-cleanup long-form name.
            normalized = normalize_venue(existing.venue, existing.year)
            if normalized and normalized != existing.venue:
                existing.venue = normalized
                existing.venue_type = detect_venue_type(existing.venue)
                changed = True
        else:
            # Cases 2 + 3: take new.
            existing.venue = new.venue
            existing.venue_source = new.venue_source
            existing.venue_type = detect_venue_type(new.venue)
            if existing.venue_type in ("conference", "workshop", "journal"):
                existing.peer_reviewed = True
            changed = True

    # Fill missing fields
    if existing.year == 0 and new.year > 0:
        existing.year = new.year
        changed = True
    if not existing.abstract and new.abstract:
        existing.abstract = new.abstract
        changed = True
    if not existing.pdf_url and new.pdf_url:
        existing.pdf_url = new.pdf_url
        changed = True
    if not existing.doi and new.doi:
        existing.doi = new.doi
        changed = True
    if not existing.arxiv_id and new.arxiv_id:
        existing.arxiv_id = new.arxiv_id
        changed = True
    if not existing.s2_paper_id and new.s2_paper_id:
        existing.s2_paper_id = new.s2_paper_id
        changed = True
    if not existing.openalex_id and new.openalex_id:
        existing.openalex_id = new.openalex_id
        changed = True

    return changed


# ---------------------------------------------------------------------------
# Write outputs
# ---------------------------------------------------------------------------

def write_outputs(papers: list[DiscoveredPaper], output_dir: Path, run: PipelineRun) -> None:
    """Write all output files: website JSON, full JSON."""
    output_dir.mkdir(parents=True, exist_ok=True)

    # Sort: year descending, then title ascending
    papers.sort(key=lambda p: (-p.year, p.title.lower()))

    # Reconcile thumbnail paths and compute stats before writing JSON
    from ndif_citations.utils import slugify

    extracted_count = 0
    missing = []

    images_dir = output_dir / "images"
    for p in papers:
        expected_path = images_dir / f"{slugify(p.title)}.png"
        if expected_path.exists():
            extracted_count += 1
            p.image = f"/images/{expected_path.name}"
        else:
            missing.append(p.title)

    run.thumbnails_extracted = extracted_count
    run.thumbnails_missing = len(missing)
    run.missing_thumbnails = missing

    from ndif_citations.models import Category
    run.low_confidence = [
        f'"{p.title}" -- classified as "{p.category.value}" (confidence: {p.category_confidence:.2f})'
        for p in papers
        if p.category_confidence < 0.7 and p.category != Category.UNCLASSIFIED
    ]

    # 1. Website JSON — verified papers only (matches ResearchPaper TS interface)
    verified_papers = [p for p in papers if p.bucket == Bucket.VERIFIED]
    website_data = [p.to_website_dict() for p in verified_papers]
    website_json = output_dir / "research-papers.json"
    with open(website_json, "w") as f:
        json.dump(website_data, f, indent=2, ensure_ascii=False)
    logger.info(f"Wrote {len(verified_papers)} verified papers to {website_json}")

    # 2. Full JSON — 3-bucket structure, all papers
    pending_papers = [p for p in papers if p.bucket == Bucket.PENDING]
    discarded_papers = [p for p in papers if p.bucket == Bucket.DISCARDED]
    full_data = {
        "pending": [p.to_full_dict() for p in pending_papers],
        "verified": [p.to_full_dict() for p in verified_papers],
        "discarded": [p.to_full_dict() for p in discarded_papers],
    }
    full_json = output_dir / "research-papers-full.json"
    with open(full_json, "w") as f:
        json.dump(full_data, f, indent=2, ensure_ascii=False, default=str)
    logger.info(
        f"Wrote {len(papers)} papers to {full_json} "
        f"({len(pending_papers)} pending, {len(verified_papers)} verified, {len(discarded_papers)} discarded)"
    )

def _write_repos_outputs(repos: list[DiscoveredRepo], output_dir: Path) -> None:
    """Write github-repos.json, github-repos-full.json."""
    # Sort: stars desc (null last), then owner/repo asc
    def _sort_key(r: DiscoveredRepo) -> tuple:
        return (r.stars is None, -(r.stars or 0), r.owner.lower(), r.repo.lower())

    repos_sorted = sorted(repos, key=_sort_key)

    # 1. Slim website JSON
    website_data = [r.to_website_dict() for r in repos_sorted]
    website_json = output_dir / "github-repos.json"
    with open(website_json, "w") as f:
        json.dump(website_data, f, indent=2, ensure_ascii=False)
    logger.info(f"Wrote {len(repos_sorted)} repos to {website_json}")

    # 2. Full state JSON
    full_data = [r.to_full_dict() for r in repos_sorted]
    full_json = output_dir / "github-repos-full.json"
    with open(full_json, "w") as f:
        json.dump(full_data, f, indent=2, ensure_ascii=False, default=str)
    logger.info(f"Wrote {len(repos_sorted)} repos to {full_json}")



def _write_repos_csv(repos: list[DiscoveredRepo], csv_path: Path) -> None:
    """Write repos to CSV."""
    fieldnames = [
        "owner", "repo", "url", "description",
        "stars", "forks", "last_commit", "archived", "is_fork",
        "language", "license", "topics",
        "repo_type", "parent_full_name",
        "category", "classification_reason",
        "linked_paper_url", "readme_arxiv_ids",
        "manual_override", "has_metadata", "has_classification",
        "content_hash", "processing_bucket",
    ]

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for repo in repos:
            writer.writerow({
                "owner": repo.owner,
                "repo": repo.repo,
                "url": repo.url,
                "description": repo.description or "",
                "stars": repo.stars if repo.stars is not None else "",
                "forks": repo.forks if repo.forks is not None else "",
                "last_commit": repo.last_commit.isoformat() if repo.last_commit else "",
                "archived": repo.archived,
                "is_fork": repo.is_fork,
                "language": repo.language or "",
                "license": repo.license or "",
                "topics": "|".join(repo.topics),
                "repo_type": repo.repo_type,
                "parent_full_name": repo.parent_full_name or "",
                "category": repo.category.value,
                "classification_reason": repo.classification_reason,
                "linked_paper_url": repo.linked_paper_url or "",
                "readme_arxiv_ids": "|".join(repo.readme_arxiv_ids),
                "manual_override": repo.manual_override,
                "has_metadata": repo.has_metadata,
                "has_classification": repo.has_classification,
                "content_hash": repo.content_hash,
                "processing_bucket": repo.processing_bucket,
            })


def _write_csv(papers: list[DiscoveredPaper], csv_path: Path) -> None:
    """Write papers to CSV with extended columns."""
    fieldnames = [
        "title", "authors", "affiliations", "venue", "year", "url", "pdf_url",
        "image", "description", "category", 
        "peer_reviewed", "venue_type", "abstract", "bibtex",
        "arxiv_id", "doi", "s2_paper_id", "openalex_id",
        "date_discovered", "category_confidence", "source",
        "manual_override", "github_repo_url",
    ]

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for paper in papers:
            row = {
                "title": paper.title,
                "authors": paper.authors,
                "affiliations": paper.affiliations,
                "venue": paper.venue,
                "year": paper.year,
                "url": paper.url,
                "pdf_url": paper.pdf_url or "",
                "image": paper.image or "",
                "description": paper.description,
                "category": paper.category.value,
                "peer_reviewed": paper.peer_reviewed,
                "venue_type": paper.venue_type or "",
                "abstract": paper.abstract or "",
                "bibtex": paper.bibtex or "",
                "arxiv_id": paper.arxiv_id or "",
                "doi": paper.doi or "",
                "s2_paper_id": paper.s2_paper_id or "",
                "openalex_id": paper.openalex_id or "",
                "date_discovered": paper.date_discovered.isoformat() if paper.date_discovered else "",
                "category_confidence": f"{paper.category_confidence:.2f}",
                "source": paper.source.value,
                "manual_override": paper.manual_override,
                "github_repo_url": paper.github_repo_url or "",
            }
            writer.writerow(row)


def _write_xlsx(
    papers: list[DiscoveredPaper],
    repos: list[DiscoveredRepo],
    output_dir: Path,
    skip_papers: bool = False,
    skip_github: bool = False,
) -> None:
    """Write research-data.xlsx with Papers and GitHub sheets."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    def _write_paper_rows(ws, paper_list: list, cols: list[str], url_cols: set[str]) -> None:
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
        for row_idx, paper in enumerate(paper_list, 2):
            row_data = {
                "title": paper.title,
                "authors": paper.authors,
                "affiliations": paper.affiliations,
                "venue": paper.venue,
                "year": paper.year,
                "url": paper.url,
                "pdf_url": paper.pdf_url or "",
                "github_repo_url": paper.github_repo_url or "",
                "description": paper.description,
                "category": paper.category.value,
                "bucket": paper.bucket.value,
                "reason": paper.reason.value if paper.reason else "",
                "reason_detail": paper.reason_detail or "",
                "peer_reviewed": paper.peer_reviewed,
                "venue_type": paper.venue_type or "",
                "arxiv_id": paper.arxiv_id or "",
                "doi": paper.doi or "",
                "date_discovered": paper.date_discovered.isoformat() if paper.date_discovered else "",
                "category_confidence": f"{paper.category_confidence:.2f}",
                "source": paper.source.value,
                "manual_override": paper.manual_override,
            }
            for col_idx, col_name in enumerate(cols, 1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_name in url_cols and isinstance(value, str) and value.startswith("http"):
                    cell.hyperlink = value
                    cell.style = "Hyperlink"

    # --- Papers sheet (verified only) ---
    if not skip_papers:
        verified_papers = [p for p in papers if p.bucket == Bucket.VERIFIED]
        ws_papers = wb.create_sheet("Papers")
        paper_cols = [
            "title", "authors", "affiliations", "venue", "year",
            "url", "pdf_url", "github_repo_url",
            "description", "category",
            "peer_reviewed", "venue_type",
            "arxiv_id", "doi", "date_discovered",
            "category_confidence", "source", "manual_override",
        ]
        url_columns = {"url", "pdf_url", "github_repo_url"}
        _write_paper_rows(ws_papers, verified_papers, paper_cols, url_columns)

        # --- Pending sheet ---
        pending_papers = [p for p in papers if p.bucket == Bucket.PENDING]
        ws_pending = wb.create_sheet("Pending")
        pending_cols = [
            "title", "authors", "venue", "year", "url",
            "category", "reason", "reason_detail",
            "arxiv_id", "doi", "source", "manual_override",
        ]
        _write_paper_rows(ws_pending, pending_papers, pending_cols, {"url"})

        # --- Discarded sheet ---
        discarded_papers = [p for p in papers if p.bucket == Bucket.DISCARDED]
        ws_discarded = wb.create_sheet("Discarded")
        discarded_cols = [
            "title", "authors", "venue", "year", "url",
            "reason", "reason_detail", "arxiv_id", "doi", "source", "manual_override",
        ]
        _write_paper_rows(ws_discarded, discarded_papers, discarded_cols, {"url"})

    # --- GitHub sheet ---
    if not skip_github:
        ws_github = wb.create_sheet("GitHub")
        github_cols = [
            "owner", "repo", "url", "description",
            "stars", "forks", "last_commit",
            "language", "license", "topics",
            "category", "classification_reason",
            "repo_type", "parent_full_name",
            "linked_paper_url",
            "manual_override", "has_metadata", "has_classification",
        ]
        # Header row
        for col_idx, col_name in enumerate(github_cols, 1):
            cell = ws_github.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)

        # Sort repos: stars desc, null last
        repos_sorted = sorted(
            repos,
            key=lambda r: (r.stars is None, -(r.stars or 0), r.owner.lower(), r.repo.lower())
        )
        github_url_cols = {"url", "linked_paper_url"}
        for row_idx, repo in enumerate(repos_sorted, 2):
            row_data = {
                "owner": repo.owner,
                "repo": repo.repo,
                "url": repo.url,
                "description": repo.description or "",
                "stars": repo.stars,
                "forks": repo.forks,
                "last_commit": repo.last_commit.isoformat() if repo.last_commit else "",
                "language": repo.language or "",
                "license": repo.license or "",
                "topics": "|".join(repo.topics),
                "category": repo.category.value,
                "classification_reason": repo.classification_reason,
                "repo_type": repo.repo_type,
                "parent_full_name": repo.parent_full_name or "",
                "linked_paper_url": repo.linked_paper_url or "",
                "manual_override": repo.manual_override,
                "has_metadata": repo.has_metadata,
                "has_classification": repo.has_classification,
            }
            for col_idx, col_name in enumerate(github_cols, 1):
                value = row_data[col_name]
                cell = ws_github.cell(row=row_idx, column=col_idx, value=value)
                if col_name in github_url_cols and isinstance(value, str) and value.startswith("http"):
                    cell.hyperlink = value
                    cell.style = "Hyperlink"

    xlsx_path = output_dir / "research-data.xlsx"
    wb.save(xlsx_path)
    logger.info(f"Wrote research-data.xlsx ({wb.sheetnames}) to {xlsx_path}")


# ---------------------------------------------------------------------------
# CLI Report
# ---------------------------------------------------------------------------

def print_report(
    run: PipelineRun,
    papers: list[DiscoveredPaper],
    output_dir: Path,
    repos: list[DiscoveredRepo] | None = None,
    skip_github: bool = False,
    skip_papers: bool = False,
    repos_removed_counts: dict[str, int] | None = None,
) -> None:
    """Print a rich CLI summary report."""
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table

    console = Console()

    # Header
    console.print()
    console.print(Panel.fit(
        "[bold cyan]NDIF Citation Tracker -- Run Complete[/bold cyan]",
        border_style="cyan",
    ))
    console.print()

    # Sources
    console.print("[bold]Sources checked:[/bold]")
    console.print(f"  > Semantic Scholar citations: [green]{run.s2_citations_found}[/green] found")
    console.print(f"  > OpenAlex fulltext search: [green]{run.openalex_found}[/green] found")
    console.print(f"  > Google Scholar (SerpAPI): [green]{run.scholar_found}[/green] found")
    console.print(f"  > GitHub dependents: [green]{run.github_dependents_found}[/green] repos with papers")
    console.print()

    # Merge results
    console.print(f"[bold]After deduplication:[/bold] [cyan]{run.total_unique}[/cyan] unique papers")
    console.print()

    if run.existing_papers > 0:
        console.print("[bold]Merge results:[/bold]")
        if run.new_papers > 0:
            console.print(f"  * [bold green]{run.new_papers} NEW[/bold green] papers added")
        else:
            console.print("  - [dim]No new papers found[/dim]")
        console.print(f"  > {run.existing_papers} already in database")
        if run.updated_papers > 0:
            console.print(f"  > {run.updated_papers} papers updated")
    else:
        console.print(f"[bold]First run:[/bold] [green]{run.total_unique}[/green] papers cataloged")
    console.print()

    # 3-bucket breakdown
    verified_papers = [p for p in papers if p.bucket == Bucket.VERIFIED]
    pending_papers = [p for p in papers if p.bucket == Bucket.PENDING]
    discarded_papers = [p for p in papers if p.bucket == Bucket.DISCARDED]

    n_ndif = sum(1 for p in verified_papers if p.category.value == "uses_ndif")
    n_nnsight = sum(1 for p in verified_papers if p.category.value == "uses_nnsight")
    n_ref = sum(1 for p in verified_papers if p.category.value == "referencing")

    def _reason_breakdown(paper_list) -> str:
        counts: dict[str, int] = {}
        for p in paper_list:
            key = p.reason.value if p.reason else "unknown"
            counts[key] = counts.get(key, 0) + 1
        return ", ".join(f"{v} {k}" for k, v in sorted(counts.items()))

    console.print("[bold]Buckets:[/bold]")
    console.print(
        f"  > [green]{len(verified_papers)} verified[/green] — "
        f"{n_ndif} uses_ndif, {n_nnsight} uses_nnsight, {n_ref} referencing"
    )
    console.print(
        f"  > [yellow]{len(pending_papers)} pending[/yellow]"
        + (f" — {_reason_breakdown(pending_papers)}" if pending_papers else "")
    )
    console.print(
        f"  > [red]{len(discarded_papers)} discarded[/red]"
        + (f" — {_reason_breakdown(discarded_papers)}" if discarded_papers else "")
    )
    console.print(f"  > {run.thumbnails_extracted} thumbnails extracted")
    if run.thumbnails_missing > 0:
        console.print(f"  ! {run.thumbnails_missing} papers need manual thumbnails")
    console.print()

    # Auto-promoted this run
    if run.auto_promoted:
        console.print(f"[bold green]Auto-promoted this run ({len(run.auto_promoted)}):[/bold green]")
        for title in run.auto_promoted[:15]:
            console.print(f'  - "{title}" (was: pending → verified)')
        if len(run.auto_promoted) > 15:
            console.print(f"  ... and {len(run.auto_promoted) - 15} more")
        console.print()

    # Auto-demoted this run
    if run.auto_demoted:
        console.print(f"[bold yellow]Auto-demoted this run ({len(run.auto_demoted)}):[/bold yellow]")
        for title in run.auto_demoted[:15]:
            console.print(f'  - "{title}" (was: verified → pending)')
        if len(run.auto_demoted) > 15:
            console.print(f"  ... and {len(run.auto_demoted) - 15} more")
        console.print()

    # Discarded breakdown
    if discarded_papers:
        console.print(f"[bold red]Discarded ({len(discarded_papers)}):[/bold red]")
        for p in discarded_papers[:15]:
            console.print(f'  - "{p.title}" ({p.reason.value if p.reason else "unknown"})')
        if len(discarded_papers) > 15:
            console.print(f"  ... and {len(discarded_papers) - 15} more")
        console.print()

    # Missing thumbnails
    if run.missing_thumbnails:
        console.print("[bold yellow]Papers needing manual thumbnails:[/bold yellow]")
        for i, title in enumerate(run.missing_thumbnails[:15], 1):
            console.print(f"  {i}. \"{title}\"")
        if len(run.missing_thumbnails) > 15:
            console.print(f"  ... and {len(run.missing_thumbnails) - 15} more")
        console.print()

    # GitHub Repos section (only when GitHub ran)
    if not skip_github and repos is not None:
        console.print()
        console.print(Panel.fit(
            "[bold green]GitHub Repos Summary[/bold green]",
            border_style="green",
        ))
        console.print()

        total_repos = len(repos)
        ndif_repos = sum(1 for r in repos if r.category.value == "uses_ndif")
        nnsight_repos = total_repos - ndif_repos

        # Bucket breakdown from processing_bucket field
        bucket_counts: dict[str, int] = {}
        for r in repos:
            bucket_counts[r.processing_bucket] = bucket_counts.get(r.processing_bucket, 0) + 1

        # Classification breakdown
        ndif_kw = sum(1 for r in repos if r.classification_reason == "ndif_keyword_match")
        gh_dep = sum(1 for r in repos if r.classification_reason == "github_dependent")

        console.print(f"  > [cyan]{total_repos}[/cyan] repos in database")
        console.print(f"  > {ndif_repos} classified as [green]\"uses_ndif\"[/green] (NDIF keyword in README)")
        console.print(f"  > {nnsight_repos} classified as [green]\"uses_nnsight\"[/green]")
        console.print()
        console.print("[bold]Routing breakdown:[/bold]")
        for bucket in ("new", "reprocess", "fill_gaps", "skip", "protected"):
            count = bucket_counts.get(bucket, 0)
            if count > 0 or bucket in ("new", "skip"):
                console.print(f"  > {count} [{bucket.upper()}]")
        if repos_removed_counts:
            total_removed = sum(repos_removed_counts.values())
            if total_removed > 0:
                console.print(f"  ! {total_removed} repo(s) removed this run:")
                for reason, count in repos_removed_counts.items():
                    if count > 0:
                        console.print(f"      - {count} via {reason}")
        console.print()
        console.print("[bold]Classification breakdown:[/bold]")
        console.print(f"  > {ndif_kw} via NDIF keyword match in README")
        console.print(f"  > {gh_dep} default (uses_nnsight, no NDIF keywords found)")
        console.print()

        # Type breakdown
        research_repos = sum(1 for r in repos if r.repo_type == "research")
        course_repos = sum(1 for r in repos if r.repo_type == "course")
        experiment_repos = sum(1 for r in repos if r.repo_type == "experiment")

        console.print("[bold]Type breakdown:[/bold]")
        console.print(f"  > {research_repos} \\[research]")
        console.print(f"  > {course_repos} \\[course]")
        console.print(f"  > {experiment_repos} \\[experiment]")
        console.print()

    # Output files
    console.print("[bold]Output files:[/bold]")
    console.print(f"  -> {output_dir / 'research-papers.json'}  ({len(verified_papers)} verified papers)")
    console.print(f"  -> {output_dir / 'research-papers-full.json'}  ({run.total_unique} total, {len(pending_papers)} pending, {len(discarded_papers)} discarded)")
    console.print(f"  -> {output_dir / 'images/'}  ({run.thumbnails_extracted} thumbnails)")
    if not skip_github:
        console.print(f"  -> {output_dir / 'github-repos.json'}  ({len(repos) if repos else 0} repos)")
        console.print(f"  -> {output_dir / 'github-repos-full.json'}  ({len(repos) if repos else 0} repos, all metadata)")
    console.print(f"  -> {output_dir / 'research-data.xlsx'}  (Papers + GitHub sheets)")
    console.print()

    # Usage hint
    console.print("[dim]To use with the NDIF website:[/dim]")
    console.print("[dim]  1. Copy output/research-papers.json -> data/research-papers.json in the website repo[/dim]")
    console.print("[dim]  2. Copy output/images/*.png -> public/images/ in the website repo[/dim]")
    console.print()

    # Errors
    if run.errors:
        console.print(f"[bold red]Errors ({len(run.errors)}):[/bold red]")
        for err in run.errors:
            console.print(f"  x {err}")
    else:
        console.print("[green]Errors (0): (none)[/green]")

    console.print()
